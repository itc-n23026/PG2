#! python3
#  updateProduce.py

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

# 農産物の種類と、更新する価格
PRICE_UPDATES = {'Garlic': 3.07, 'Glery': 1.19, 'Lemon': 1.27}

# 行をループして価格を更新する
for row_num in range(2, sheet.max_row + 1):
    produce_name = sheet.cell(row=row_num, column=1).value
    if produce_name in PRICE_UPDATES:
        sheet.cell(row=row_num, column=2).value = PRICE_UPDATES[produce_name]

wb.save('updateProduceSales.xlsx')