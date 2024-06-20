import sys
import openpyxl

def transpose_sheet(input_filename, output_filename):
    # 入力ファイルを開く
    wb = openpyxl.load_workbook(input_filename)
    ws = wb.active
    
    # 新しいワークブックとワークシートを作成
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    
    # 行と列を入れ替えて新しいワークシートに書き込む
    rows = list(ws.iter_rows(values_only=True))
    for row_idx, row in enumerate(rows):
        for col_idx, value in enumerate(row):
            new_ws.cell(row=col_idx + 1, column=row_idx + 1, value=value)
    
    # 新しいファイルに保存
    new_wb.save(output_filename)
    print(f"{output_filename} が作成されました。")

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("使用方法: python3 swaprowcol.py <入力ファイル名> <出力ファイル名>")
        sys.exit(1)
    
    try:
        input_filename = sys.argv[1]
        output_filename = sys.argv[2]
        transpose_sheet(input_filename, output_filename)
    except FileNotFoundError:
        print(f"エラー: ファイル {input_filename} が見つかりません。")
        sys.exit(1)
    except Exception as e:
        print(f"エラーが発生しました: {e}")
        sys.exit(1)